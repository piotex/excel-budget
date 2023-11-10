import csv
import re
from models.CsvData import CsvData
import os
from typing import List
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime


def get_list_of_priv_files():
    csv_pattern = re.compile(r'\.csv$', re.IGNORECASE)
    main_path = os.path.abspath(__file__)
    main_directory = os.path.dirname(main_path)
    ala_folder_path = os.path.join(main_directory, "priv-data")

    file_list = [
        os.path.join(ala_folder_path, file)
        for file in os.listdir(ala_folder_path)
        if os.path.isfile(os.path.join(ala_folder_path, file)) and csv_pattern.search(file)
    ]
    return file_list


def read_csv_to_dataclass_list(csv_file_path, encoding='utf-8'):
    data_objects = []
    with open(csv_file_path, 'r', encoding=encoding) as csvfile:
        csvreader = csv.DictReader(csvfile, delimiter=';')
        for row in csvreader:
            model = CsvData(
                Data_księgowania=row['Data księgowania'],
                Data_waluty=row['Data waluty'],
                Nadawca_Odbiorca=row['Nadawca / Odbiorca'],
                Adres_nadawcy_odbiorcy=row['Adres nadawcy / odbiorcy'],
                Rachunek_źródłowy=row['Rachunek źródłowy'],
                Rachunek_docelowy=row['Rachunek docelowy'],
                Tytułem=row['Tytułem'],
                Kwota_operacji=float(row['Kwota operacji'].replace(',', '.').replace(' ', '')),
                Waluta=row['Waluta'],
                Numer_referencyjny=row['Numer referencyjny'],
                Typ_operacji=row['Typ operacji']
            )
            model.add_my_category()
            data_objects.append(model)
    return data_objects


def save_data_to_sheet(file_path, sheet_name: str, list_of_data: List[CsvData]):
    book = load_workbook(file_path)
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        book.remove(sheet)
        book.save(file_path)

        df = pd.DataFrame([ob.__dict__ for ob in list_of_data])
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
    else:
        raise Exception(f"=== my Exception === No {sheet_name} in Excel file.")


def save_list_of_unique_csv_data(list_of_unique_csv_data: List[CsvData]):
    actual_date = datetime.now().strftime('%Y.%m.%d')
    excel_file_path = f"priv-data/excel-bank-{actual_date}.xlsx"

    try:
        save_data_to_sheet(excel_file_path, 'all_data', list_of_unique_csv_data)

        filtr = lambda obiekt: obiekt.my_category == 'savings'
        list_of_savings = list(filter(filtr, list_of_unique_csv_data))
        for elem in list_of_savings:
            elem.Kwota_operacji = abs(elem.Kwota_operacji)
        save_data_to_sheet(excel_file_path, 'savings', list_of_savings)

        filtr = lambda obiekt: obiekt.my_category == 'income'
        list_of_savings = list(filter(filtr, list_of_unique_csv_data))
        save_data_to_sheet(excel_file_path, 'income', list_of_savings)

        filtr = lambda obiekt: obiekt.my_category == 'expense'
        list_of_savings = list(filter(filtr, list_of_unique_csv_data))
        save_data_to_sheet(excel_file_path, 'expense', list_of_savings)
    except FileNotFoundError:
        raise Exception(f"=== my Exception === No {excel_file_path} Excel file.")
